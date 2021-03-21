# https://www.kaggle.com/shivamb/netflix-shows
import argparse
from os import path, getcwd
import sys
import csv
from openpyxl.styles import Color, Font
import copy
import openpyxl

# task 7
parser = argparse.ArgumentParser(description="""The application is meant
                                to proces netflix-shows data and
                                performs some statistical tasks on them""")
parser.add_argument('dataset', help='a csv file name to analyzed', type=str)
parser.add_argument('-o', '--out', help='output file excel forma and '
                    'extenstion', type=str)
args = parser.parse_args()
# get the current working directoy


def read_csv(dataset_file):
    """ Read the data set and return an object dictionary of it"""
    list_csv = []
    try:
        # read the csv
        with open(dataset_file, newline='',
                  encoding='utf-8') as infile_dataset:
            # read as csv
            dict_reader = csv.DictReader(infile_dataset)
            for row in dict_reader:
                realeased_on = 0
                if row['release_year'].isdigit():
                    realeased_on = int(row['release_year'])
                new_dict = {
                    'show_id': row['show_id'], 'type': row['type'],
                    'title': row['title'], 'director': row['director'],
                    'cast': row['cast'], 'country': row['country'],
                    'date_added': row['date_added'],
                    'release_year': realeased_on,
                    'rating': row['rating'], 'duration': row['duration'],
                    'listed_in': row['listed_in'],
                    'description': row['description']
                }
                list_csv.append(new_dict)
    except OSError:
        print('Could not open the file wrong file or/ resources not available')
        sys.exit()
    return list_csv


# Task 4
def get_latest_movie(list_dataset):
    """Get the list of dictionary of data set and returns the latest movie"""
    latest_movie_data = max(list_dataset, key=lambda dic: dic['release_year'])
    return latest_movie_data


def get_first_movie(list_dataset):
    """Get the list of dict and returns the earliest released movie"""
    latest_movie_data = min(list_dataset, key=lambda dic: dic['release_year'])
    return latest_movie_data


def get_country_movies(list_dataset):
    """Returns a list countries and the number of movies per countries"""
    country_dict = {}
    for film in list_dataset:
        country_name = film.get('country')
        if country_name in country_dict:
            country_dict[country_name] = country_dict[country_name] + 1
        else:
            country_dict[country_name] = 1
    return country_dict


def get_country_with_max_production(list_dataset):
    """Get the list of dataset and return the country
    name with the highest production number"""
    production_per_country = get_country_movies(list_dataset)
    country_name = max(production_per_country,
                       key=lambda k: production_per_country[k])
    return {country_name: production_per_country[country_name]}


def get_total_number_movies(list_dataset):
    """Returns the total numbers of movies product"""
    # because each movie has 1 unique if and different
    # therefore there is no repetition
    return len(list_dataset)


def get_movies_number_per_year(list_dataset):
    """Returns the number of movies per years"""
    year_dict = {}
    for film in list_dataset:
        movie_year = film.get('release_year')
        if movie_year in year_dict:
            year_dict[movie_year] = year_dict[movie_year] + 1
        else:
            year_dict[movie_year] = 1
    return year_dict


def get_most_productive_year(list_dataset):
    """Get the dataset list and returns the most productive year"""
    year_dict = get_movies_number_per_year(list_dataset)
    year = max(year_dict, key=lambda k: year_dict[k])
    return year


# Task 5, 6
def analize_data(data, out_file):
    """Analyze the data set and display the result"""
    latest_movie = get_latest_movie(data)
    first_movie = get_first_movie(data)
    number_of_movies_per_country = get_country_movies(data)
    country_with_max_production = get_country_with_max_production(data)
    realse_per_year = get_movies_number_per_year(data)
    most_productive_year = get_most_productive_year(data)
    total_number_movies = get_total_number_movies(data)
    if out_file is None or not out_file.endswith('.xlsx'):
        print("The latest movie released")
        for key in latest_movie:
            print("{} : {} ".format(key, latest_movie[key]), end=' ')
        print("\nThe first movie released")
        for key in first_movie:
            print("{} : {} ".format(key, latest_movie[key]), end=' ')
        print("\nContry name (group)", "Number of movies")
        for key in number_of_movies_per_country:
            print("{} : {} ".format(key, number_of_movies_per_country[key]))
        print('Country with max movie production')
        print(country_with_max_production)
        print("Number of movieds released per year")
        for key in realse_per_year:
            print("{} : {} ".format(key, realse_per_year[key]))
        print("The most productive year")
        print(most_productive_year)
        print("The total number of movies")
        print(total_number_movies)
    else:
        # create a workbook
        wb = openpyxl.Workbook()
        # simple statistical data
        sheet_1 = wb.active
        sheet_1.title = "Simple statistical information"
        sheet_2 = wb.create_sheet(title="Number of movies per country")
        sheet_3 = wb.create_sheet(title="Number of movies per year")
        # latest released movie
        ft = Font(color="00333399", bold=True, size=12, )
        sheet_1.append(["The latest movie released"])
        sheet_1.append(list(latest_movie.keys()))
        sheet_1.append(list(latest_movie.values()))
        # first movie released
        sheet_1.append(["The first  movie released"])
        sheet_1.append(list(first_movie.keys()))
        sheet_1.append(list(first_movie.values()))
        # country with max production
        sheet_1.append(["The counrty with the most number of movies"])
        country_name = list(country_with_max_production.keys())[0]
        country_number = country_with_max_production[country_name]
        sheet_1.append([country_name, country_number])
        # The most productive year
        sheet_1.append(["The most productive country"])
        sheet_1.append(["The most productive year : ", most_productive_year])
        sheet_1.append(["The most total number of produced movies"])
        sheet_1.append(["The total number of movies produced : ",
                        total_number_movies])
        # simple styles
        b8 = sheet_1['B8']
        b10 = sheet_1['B10']
        b12 = sheet_1['B12']
        b8.font = ft
        b10.font = ft
        b12.font = ft
        # number of movies per country
        sheet_2.append(["Contry name (group)", "Number of movies"])
        count = 2
        for country_key in number_of_movies_per_country:
            sheet_2.append([country_key,
                            number_of_movies_per_country[country_key]])
            x = sheet_2['B{}'.format(count)]
            x.font = ft
            count += 1
        # number of movies per year
        sheet_3.append(["Year", "Number of movies"])
        for country_key in realse_per_year:
            sheet_3.append([country_key,
                            realse_per_year[country_key]])

        wb.save(out_file)


# Task 1 , 2
def main():
    """Main function of my app"""
    file_name = args.dataset
    out_file = args.out
    if not file_name.endswith('.csv'):
        print('The provided file name is not a csv extension')
        sys.exit()
    # get the current working directory that works in all OS
    file_path = path.join(getcwd(), file_name)
    if not path.isfile(file_path):
        print('The provided file name does not exists')
        sys.exit()
    data = read_csv(file_name)
    analize_data(data, out_file)


if __name__ == '__main__':
    main()
